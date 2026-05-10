import json
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


ROOT = Path(__file__).resolve().parent
JEST_JSON = ROOT / "tmp-jest-results.json"
OUTPUT_XLSX = ROOT / "tests" / "TestResult_ByFunction.xlsx"
TEST_FILE = ROOT / "tests" / "exam.service.test.ts"


DETAIL = {
    "TC_ES_UPD_006": {
        "goal": "Validate Title không rỗng khi update",
        "input": 'updateData:{Title:""}',
        "expected": 'Throw "Exam title cannot be empty"',
    },
    "TC_ES_UPD_007": {
        "goal": "Validate Title không chỉ whitespace",
        "input": 'updateData:{Title:"   "}',
        "expected": 'Throw "Exam title cannot be empty"',
    },
    "TC_ES_AQ_006": {
        "goal": "Chặn payload questions rỗng",
        "input": "questions:[]",
        "expected": 'Throw "Questions payload cannot be empty"',
    },
    "TC_ES_AQ_007": {
        "goal": "Chặn duplicate QuestionID trong request addQuestions",
        "input": "questions:[{QuestionID:1,OrderIndex:1},{QuestionID:1,OrderIndex:2}]",
        "expected": 'Throw "Duplicate QuestionID in request payload"',
    },
    "TC_ES_AQ_008": {
        "goal": "Chặn duplicate OrderIndex trong request addQuestions",
        "input": "questions:[{QuestionID:1,OrderIndex:10},{QuestionID:2,OrderIndex:10}]",
        "expected": 'Throw "Duplicate OrderIndex in request payload"',
    },
    "TC_ES_RQ_004": {
        "goal": "Chặn questionIds rỗng khi remove",
        "input": "questionIds:[]",
        "expected": 'Throw "Question IDs cannot be empty"',
    },
    "TC_ES_MMG_002": {
        "goal": "Validate newStartOrderIndex > 0",
        "input": "newStartOrderIndex=0",
        "expected": 'Throw "newStartOrderIndex must be greater than 0"',
    },
    "TC_ES_CRE_013": {
        "goal": "Chặn duplicate QuestionID trong payload",
        "input": "questions:[{QuestionID:1,OrderIndex:1},{QuestionID:1,OrderIndex:2}]",
        "expected": 'Throw "Duplicate QuestionID in request payload"',
    },
    "TC_ES_CRE_014": {
        "goal": "Chặn duplicate OrderIndex trong payload",
        "input": "questions:[{QuestionID:1,OrderIndex:1},{QuestionID:2,OrderIndex:1}]",
        "expected": 'Throw "Duplicate OrderIndex in request payload"',
    },
    "TC_ES_CRE_015": {
        "goal": "Validate QuestionID > 0",
        "input": "questions:[{QuestionID:0,OrderIndex:1}]",
        "expected": 'Throw "QuestionID must be a positive integer"',
    },
    "TC_ES_CRE_016": {
        "goal": "Validate MediaQuestionID > 0",
        "input": "MediaQuestionIDs:[0]",
        "expected": 'Throw "MediaQuestionID must be a positive integer"',
    },
    "TC_ES_CRE_017": {
        "goal": "Validate userId > 0",
        "input": "userId=0",
        "expected": 'Throw "UserID must be a positive integer"',
    },
    "TC_ES_GET_004": {
        "goal": "Fail graceful khi thiếu examType relation",
        "input": "exam.examType=null",
        "expected": 'Throw "Exam type data is missing"',
    },
    "TC_ES_GET_005": {
        "goal": "Fail graceful khi question thiếu media relation",
        "input": "question.mediaQuestion=null",
        "expected": 'Throw "Question media data is missing"',
    },
    "TC_ES_GCO_002": {
        "goal": "Fail graceful khi standalone question thiếu media",
        "input": "standalone.question.mediaQuestion=null",
        "expected": 'Throw "Standalone question media data is missing"',
    },
    "TC_ES_RMG_005": {
        "goal": "Validate mediaQuestionId > 0",
        "input": "mediaQuestionId=0",
        "expected": 'Throw "MediaQuestionID must be a positive integer"',
    },
}


def split_tc_title(title: str) -> tuple[str, str]:
    if ": " in title and title.startswith("TC_"):
        tcid, usecase = title.split(": ", 1)
        return tcid.strip(), usecase.strip()
    return "", title


def parse_method(ancestor_titles: list[str]) -> str:
    if len(ancestor_titles) >= 2:
        return ancestor_titles[1].split(" - ")[0].strip()
    return ""


def method_from_id(tcid: str, title: str = "") -> str:
    if not tcid:
        return ""
    prefix_map = [
        ("_CRE_", "createExam"),
        ("_GET_", "getExamById"),
        ("_GALL_", "getAllExams"),
        ("_UPD_", "updateExam"),
        ("_DEL_", "deleteExam"),
        ("_AQ_", "addQuestionsToExam"),
        ("_RQ_", "removeQuestionsFromExam"),
        ("_STAT_", "getExamStatistics"),
        ("_SRCH_", "searchExams"),
        ("_DUP_", "duplicateExam"),
        ("_AMG_", "addMediaGroupToExam"),
        ("_RMG_", "removeMediaGroupFromExam"),
        ("_ET_", "ExamType CRUD"),
        ("_NOI_", "getNextOrderIndex"),
        ("_VES_", "validateExamStructure"),
        ("_CEO_", "compactExamOrder"),
        ("_RPQ_", "replaceQuestionInExam"),
        ("_MMG_", "moveMediaGroupInExam"),
        ("_MGS_", "getExamMediaGroupSummary"),
        ("_GCO_", "getExamContentOrganized"),
    ]
    for token, method in prefix_map:
        if token in tcid:
            return method

    lower = title.lower()
    if "movemediagroupinexam" in lower:
        return "moveMediaGroupInExam"
    if "getexammediagroupsummary" in lower:
        return "getExamMediaGroupSummary"
    if "getexamcontentorganized" in lower:
        return "getExamContentOrganized"
    if "updateexam" in lower:
        return "updateExam"
    if "addquestionstoexam" in lower:
        return "addQuestionsToExam"
    if "removequestionsfromexam" in lower:
        return "removeQuestionsFromExam"
    return ""


def function_group(method: str) -> str:
    m = method.lower()
    if "createexam" in m:
        return "Tạo bài test"
    if "getexambyid" in m:
        return "Lấy thông tin bài test"
    if "getallexams" in m:
        return "Danh sách bài test"
    if "updateexam" in m:
        return "Cập nhật bài test"
    if "deleteexam" in m:
        return "Xóa bài test"
    if "addquestionstoexam" in m:
        return "Thêm câu hỏi vào bài test"
    if "removequestionsfromexam" in m:
        return "Xóa câu hỏi khỏi bài test"
    if "getexamstatistics" in m:
        return "Thống kê bài test"
    if "searchexams" in m:
        return "Tìm kiếm bài test"
    if "duplicateexam" in m:
        return "Nhân bản bài test"
    if "addmediagrouptoexam" in m:
        return "Thêm media group"
    if "removemediagroupfromexam" in m:
        return "Xóa media group"
    if "getnextorderindex" in m:
        return "Lấy thứ tự câu hỏi"
    if "validateexamstructure" in m:
        return "Validate cấu trúc đề"
    if "compactexamorder" in m:
        return "Nén thứ tự câu hỏi"
    if "replacequestioninexam" in m:
        return "Thay thế câu hỏi"
    if "movemediagroupinexam" in m:
        return "Di chuyển media group"
    if "getexammediagroupsummary" in m:
        return "Tổng hợp media group"
    if "getexamcontentorganized" in m:
        return "Nội dung đề theo nhóm"
    if "examtype" in m:
        return "Quản lý loại đề"
    return "Quản lý bài test đầu vào"


def parse_test_doc_map() -> dict[str, dict[str, str]]:
    """
    Parse block comments immediately above each testcase to extract:
    - usecase title
    - objective
    - input
    - expected
    - notes
    """
    if not TEST_FILE.exists():
        return {}

    content = TEST_FILE.read_text(encoding="utf-8", errors="ignore")
    pattern = re.compile(
        r"/\*\*\s*"
        r"(?:\n\s*\*\s*(TC_[A-Z0-9_]+):\s*(.+?)\s*)"
        r"(?:\n\s*\*\s*)+Test Objective:\s*(.+?)\s*"
        r"(?:\n\s*\*\s*)+Input:\s*(.+?)\s*"
        r"(?:\n\s*\*\s*)+Expected Output:\s*(.+?)\s*"
        r"(?:\n\s*\*\s*Notes:\s*(.+?)\s*)?"
        r"\n\s*\*/",
        re.DOTALL,
    )
    doc_map: dict[str, dict[str, str]] = {}
    for m in pattern.finditer(content):
        tcid = m.group(1).strip()
        usecase = m.group(2).strip()
        objective = m.group(3).strip()
        input_txt = m.group(4).strip()
        expected_txt = m.group(5).strip()
        notes_txt = (m.group(6) or "").strip()
        doc_map[tcid] = {
            "usecase": re.sub(r"\s+", " ", usecase),
            "objective": re.sub(r"\s+", " ", objective),
            "input": re.sub(r"\s+", " ", input_txt),
            "expected": re.sub(r"\s+", " ", expected_txt),
            "notes": re.sub(r"\s+", " ", notes_txt),
        }
    return doc_map


def build_rows(data: dict[str, Any]) -> list[dict[str, str]]:
    doc_map = parse_test_doc_map()
    rows: list[dict] = []
    for suite in data.get("testResults", []):
        for assertion in suite.get("assertionResults", []):
            title = assertion.get("title", "")
            tcid, usecase = split_tc_title(title)
            method = (
                method_from_id(tcid, usecase)
                or parse_method(assertion.get("ancestorTitles", []))
            )
            usecase_group = function_group(method)
            status = "PASS" if assertion.get("status") == "passed" else "FAIL"
            default_goal = usecase
            default_input = "N/A"
            default_expected = "N/A"
            if tcid in doc_map:
                default_goal = doc_map[tcid]["objective"] or default_goal
                default_input = doc_map[tcid]["input"] or default_input
                default_expected = doc_map[tcid]["expected"] or default_expected

            if tcid in DETAIL:
                default_goal = DETAIL[tcid]["goal"]
                default_input = DETAIL[tcid]["input"]
                default_expected = DETAIL[tcid]["expected"]

            note = ""
            if status == "FAIL" and assertion.get("failureMessages"):
                note = " ".join(assertion["failureMessages"]).replace("\n", " ")
                if len(note) > 500:
                    note = note[:500] + "..."
            elif tcid in doc_map and doc_map[tcid].get("notes"):
                note = doc_map[tcid]["notes"]

            rows.append(
                {
                    "TestcaseID": tcid,
                    "Chức năng/use case": usecase_group,
                    "Lớp": "ExamService",
                    "Phương thức": method,
                    "Mục tiêu kiểm thử": default_goal,
                    "Input (Dữ liệu mock)": default_input,
                    "Expected output": default_expected,
                    "Kết quả": status,
                    "Ghi chú": note,
                }
            )
    return rows


def export_excel(rows: list[dict[str, str]], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "ExamService Tests"
    headers = [
        "TestcaseID",
        "Chức năng/use case",
        "Lớp",
        "Phương thức",
        "Mục tiêu kiểm thử",
        "Input (Dữ liệu mock)",
        "Expected output",
        "Kết quả",
        "Ghi chú",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])

    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 18,
        "B": 40,
        "C": 15,
        "D": 20,
        "E": 35,
        "F": 40,
        "G": 35,
        "H": 10,
        "I": 80,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=9):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    wb.save(out_path)


def main() -> None:
    if not JEST_JSON.exists():
        raise FileNotFoundError(
            f"Không tìm thấy {JEST_JSON.name}. Hãy chạy Jest với --json --outputFile trước."
        )
    data = json.loads(JEST_JSON.read_text(encoding="utf-8"))
    rows = build_rows(data)
    out_path = OUTPUT_XLSX
    # Fallback nếu file đích đang mở trong Excel
    try:
        export_excel(rows, out_path)
    except PermissionError:
        out_path = ROOT / "tests" / "TestResult_ByFunction_v2.xlsx"
        export_excel(rows, out_path)
        print("File đích đang bận, đã xuất sang bản v2.")
    print(f"Exported: {out_path}")
    print(f"Total test rows: {len(rows)}")


if __name__ == "__main__":
    main()
